import streamlit as st
import pandas as pd
import time
import pulp
from datetime import datetime
import openpyxl
from collections import defaultdict
import io

# ベースパターンの設定
BASE_PATTERNS = {
    'D10': [4000, 4500, 5500, 6000],
    'D13': [3000, 4000, 4500, 5500, 6000, 7500],
    'D16': [4000, 4500, 5500, 6000, 7000],
    'D19': [3500, 4000, 4500, 5500, 6000],
    'D22': [4000, 4500, 5500, 6000]
}

def find_cell_position(worksheet, search_text):
    """ワークシート内で指定されたテキストを含むセルの位置を検索"""
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value and search_text in str(cell.value):
                return cell.row, cell.column
    return None

def get_diameter_column_index(worksheet, base_row, base_col, target_diameter):
    """指定された径の列インデックスを取得"""
    col = base_col + 1
    max_col = worksheet.max_column
    
    while col <= max_col:
        cell_value = worksheet.cell(row=base_row, column=col).value
        if cell_value and str(cell_value).strip() == target_diameter:
            return col
        col += 1
    
    return None

def extract_cutting_data_from_sheet(worksheet, target_diameter):
    """単一シートから指定された径の切断データを抽出"""
    base_position = find_cell_position(worksheet, "鉄筋径")
    if not base_position:
        return {}
    
    base_row, base_col = base_position
    diameter_col = get_diameter_column_index(worksheet, base_row, base_col, target_diameter)
    if not diameter_col:
        return {}
    
    cutting_data = {}
    row = base_row + 1
    max_row = worksheet.max_row
    
    while row <= max_row:
        length_cell = worksheet.cell(row=row, column=base_col+1)
        count_cell = worksheet.cell(row=row, column=diameter_col)
        
        if (length_cell.value is not None and count_cell.value is not None and
            isinstance(length_cell.value, (int, float)) and 
            isinstance(count_cell.value, (int, float)) and
            count_cell.value > 0):
            
            length = int(length_cell.value)
            count = int(count_cell.value)
            cutting_data[length] = count
        
        if length_cell.value is None or length_cell.value == '':
            empty_count = 0
            temp_row = row
            while temp_row <= min(row + 2, max_row):
                if worksheet.cell(row=temp_row, column=base_col).value is None:
                    empty_count += 1
                temp_row += 1
            if empty_count >= 3:
                break
        
        row += 1
    
    return cutting_data

def create_result_sheets(df_results, target_diameter, uploaded_file=None):
    """結果シートを作成する関数"""
    # pattern列の最大項目数を計算
    max_items = 0
    for pattern in df_results['pattern']:
        if pd.notna(pattern):
            items = [item.strip().replace('"', '') for item in str(pattern).split(',')]
            max_items = max(max_items, len(items))

    # DataFrameをコピーして加工
    df = df_results.copy()
    
    # pattern列を展開するための新しい列を作成
    for i in range(1, max_items + 1):
        df[f'item_{i}'] = ''

    # 各行のpattern列を展開
    for index, row in df.iterrows():
        pattern = row['pattern']
        if pd.notna(pattern):
            items = [item.strip().replace('"', '') for item in str(pattern).split(',')]
            for i, item in enumerate(items):
                df.at[index, f'item_{i + 1}'] = item

    # 元のpattern列を削除
    df = df.drop('pattern', axis=1)

    # item列から全てのユニークな値を取得
    unique_values = set()
    for i in range(1, max_items + 1):
        col_name = f'item_{i}'
        for value in df[col_name]:
            if value and value.strip():
                unique_values.add(value.strip())

    # ユニークな値を降順でソート
    unique_values = sorted(list(unique_values), key=lambda x: int(x) if x.isdigit() else float('inf'), reverse=True)

    # カウント表を作成
    count_data = []
    for index, row in df.iterrows():
        id_value = row['id']
        times_value = int(row['times']) if pd.notna(row['times']) else 0
        
        count_row = {'id': id_value}
        for value in unique_values:
            count_row[value] = 0
        
        for i in range(1, max_items + 1):
            col_name = f'item_{i}'
            item_value = row[col_name]
            if item_value and item_value.strip():
                item_value = item_value.strip()
                if item_value in count_row:
                    count_row[item_value] += times_value
        
        count_data.append(count_row)

    # カウント結果をDataFrameに変換
    count_df = pd.DataFrame(count_data)

    # 合計行は空行として追加(後でExcel関数を設定)
    cutting_total_row = {'id': '合計'}
    for value in unique_values:
        cutting_total_row[str(value)] = None  # Excelでの数式設定用
    count_df = pd.concat([count_df, pd.DataFrame([cutting_total_row])], ignore_index=True)

    # 切断集計表データの処理(アップロードされたファイルがある場合)
    cutting_count_df = None
    sheet_cutting_data_all = None
    cutting_unique_values = None
    
    if uploaded_file:
        try:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            
            # 各シートから切断データを収集
            sheet_cutting_data_all = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # シートから切断データを抽出
                sheet_cutting_data = extract_cutting_data_from_sheet(worksheet, target_diameter)
                
                if sheet_cutting_data:
                    sheet_cutting_data_all[sheet_name] = sheet_cutting_data
            
            workbook.close()
            
            if sheet_cutting_data_all:
                # 切断集計表データから全てのユニークな値を取得
                cutting_unique_values = set()
                for sheet_data in sheet_cutting_data_all.values():
                    cutting_unique_values.update(sheet_data.keys())
                
                # ユニークな値を降順でソート
                cutting_unique_values = sorted(list(cutting_unique_values), 
                                             key=lambda x: int(x) if str(x).isdigit() else float('inf'), reverse=True)
                
                # 切断集計表のカウント表を作成
                cutting_count_data = []
                for sheet_name, sheet_data in sheet_cutting_data_all.items():
                    # 各ユニーク値に対するカウントを初期化
                    count_row = {'シート名': sheet_name}
                    for value in cutting_unique_values:
                        count_row[str(value)] = sheet_data.get(value, 0)
                    
                    cutting_count_data.append(count_row)
                
                # 切断集計表のカウント結果をDataFrameに変換
                cutting_count_df = pd.DataFrame(cutting_count_data)
                
                # 切断集計表カウント表に合計行を追加(後でExcel関数を設定)
                project_total_row = {'シート名': '合計'}
                for value in cutting_unique_values:
                    project_total_row[str(value)] = None  # Excelでの数式設定用
                cutting_count_df = pd.concat([cutting_count_df, pd.DataFrame([project_total_row])], ignore_index=True)
        
        except Exception as e:
            st.warning(f"切断集計表の処理中にエラーが発生しました: {e}")

    return df, count_df, cutting_count_df, sheet_cutting_data_all, cutting_unique_values

def read_cutting_data_from_xlsx(uploaded_file, target_diameter):
    """XLSXファイルから指定された径の切断データを読み込み"""
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        total_cutting_data = defaultdict(int)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_cutting_data = extract_cutting_data_from_sheet(worksheet, target_diameter)

            for length, count in sheet_cutting_data.items():
                total_cutting_data[length] += count

        workbook.close()
        return dict(total_cutting_data)

    except Exception as e:
        st.error(f"XLSXファイルの読み込み中にエラーが発生しました: {e}")
        return {}

def read_scrap_data_from_csv(uploaded_file):
    """CSVファイルから再利用端材データを読み込み、同じ長さの端材をまとめる"""
    try:
        df = pd.read_csv(uploaded_file)

        # 列名を標準化（スペースを削除）
        df.columns = df.columns.str.strip()

        # 最初の2列を使用（列名にかかわらず）
        if len(df.columns) < 2:
            st.error("CSVファイルには少なくとも2列（長さ、本数）が必要です")
            return {}

        # データを辞書に変換し、同じ長さの本数をまとめる
        scrap_data = defaultdict(int)
        for _, row in df.iterrows():
            length = row.iloc[0]
            count = row.iloc[1]

            # 数値チェック
            if pd.notna(length) and pd.notna(count):
                try:
                    length_int = int(float(length))
                    count_int = int(float(count))
                    if length_int > 0 and count_int > 0:
                        scrap_data[length_int] += count_int
                except (ValueError, TypeError):
                    continue

        return dict(scrap_data)

    except Exception as e:
        st.error(f"CSVファイルの読み込み中にエラーが発生しました: {e}")
        return {}

def dfs(index, current_combination, current_sum, remaining_counts, sorted_numbers, max_sum, all_combinations):
    """深さ優先探索の再帰関数"""
    if 0 < current_sum <= max_sum:
        sorted_combo = tuple(sorted(current_combination, reverse=True))
        all_combinations.add((sorted_combo, current_sum))
    
    if current_sum > max_sum or index >= len(sorted_numbers):
        return
    
    current_number = sorted_numbers[index]
    max_count = remaining_counts[current_number]
    
    for count in range(max_count + 1):
        new_sum = current_sum + current_number * count
        
        if new_sum > max_sum:
            break
        
        new_combination = current_combination + [current_number] * count
        new_remaining = remaining_counts.copy()
        new_remaining[current_number] -= count
        
        dfs(index + 1, new_combination, new_sum, new_remaining, sorted_numbers, max_sum, all_combinations)

def find_combinations_dfs(numbers_dict, max_sum=7500):
    """深さ優先探索を使って組み合わせを求める"""
    sorted_numbers = sorted(numbers_dict.keys(), reverse=True)
    all_combinations = set()
    
    dfs(0, [], 0, numbers_dict.copy(), sorted_numbers, max_sum, all_combinations)
    
    return [(list(combo), sum_val) for combo, sum_val in all_combinations]

def generate_all_combinations(available_rods, required_cuts):
    """指定された棒の長さから切り出し可能な全ての組み合わせを生成"""
    combinations = find_combinations_dfs(required_cuts, max(available_rods))
    combinations.sort(key=lambda x: x[1])

    available_rods.sort()
    i = 0

    all_combinations = []
    for combo, total_cut_length in combinations:
        while available_rods[i] < total_cut_length:
            i += 1
        loss = (available_rods[i] - total_cut_length) / available_rods[i]
        all_combinations.append({
            'rod_length': available_rods[i],
            'cuts': tuple(combo),
            'loss': loss,
        })
    
    all_combinations.sort(key=lambda x: x['cuts'], reverse=True)
    all_combinations.sort(key=lambda x: x['rod_length'], reverse=False)
    return all_combinations

def optimal_cutting_plan(c, a, q, time_limit=120, reuse_pattern_counts=None):
    """最適な切り出しプランを計算"""
    n = len(a)
    m = len(q)

    prob = pulp.LpProblem("CuttingStockProblem", pulp.LpMinimize)
    x = [pulp.LpVariable(f"y{j+1}", lowBound=0, cat='Integer') for j in range(n)]

    # 再利用端材の本数制約を追加
    if reuse_pattern_counts:
        pattern_index = 0
        for reuse_info in reuse_pattern_counts:
            pattern_count = reuse_info['pattern_count']
            rod_count = reuse_info['count']

            # このreuse_rodのパターンに対する制約を追加
            # x[pattern_index] + x[pattern_index+1] + ... + x[pattern_index+pattern_count-1] <= rod_count
            if pattern_count > 0:
                reuse_constraint = pulp.lpSum(x[j] for j in range(pattern_index, pattern_index + pattern_count))
                prob += reuse_constraint <= rod_count, f"Reuse_constraint_{pattern_index}"

            pattern_index += pattern_count

    objective = pulp.lpSum(c[j] * x[j] for j in range(n))
    prob += objective, "Total_Loss"

    for i in range(m):
        production_constraint = pulp.lpSum(a[j][i] * x[j] for j in range(n))
        prob += production_constraint == q[i], f"Demand_constraint_{i+1}"

    prob.solve(pulp.PULP_CBC_CMD(timeLimit=time_limit, msg=False))

    if prob.status == pulp.LpStatusOptimal:
        optimal_solution = [var.varValue for var in x]
        optimal_value = pulp.value(prob.objective)
        return optimal_solution, optimal_value
    else:
        return None, None

def display_cutting_patterns(cutting_patterns):
    """切断パターンを見やすい箇条書き形式で表示"""
    if not cutting_patterns:
        st.write("切断パターンがありません")
        return

    st.write("**最適切断指示:**")
    for i, pattern in enumerate(cutting_patterns):
        cuts_str = " + ".join([str(cut) for cut in pattern['cuts']])
        st.write(f"**{i+1}:** {pattern['rod_length']}mm → ({cuts_str}) [{pattern['loss']}] * {pattern['num']}")


def recalculate_with_threshold(cutting_patterns, scrap_threshold):
    """閾値を変更して歩留り率と端材を再計算"""

    total_rod_length = 0
    used_length = 0

    for pattern in cutting_patterns:
        total_rod_length += pattern["rod_length"] * pattern["num"]
        used_length += sum(pattern["cuts"]) * pattern["num"]

    loss = total_rod_length - used_length
    yield_rate = used_length * 100 / total_rod_length if total_rod_length > 0 else 0

    # 端材を閾値で分類
    scrap_below_threshold = 0
    scrap_above_threshold = []

    for pattern in cutting_patterns:
        pattern_loss = pattern["loss"]
        pattern_num = pattern["num"]
        if pattern_loss < scrap_threshold:
            scrap_below_threshold += pattern_loss * pattern_num
        else:
            scrap_above_threshold.append(
                {
                    "length": pattern_loss,
                    "count": pattern_num,
                    "rod_length": pattern["rod_length"],
                }
            )

    yield_rate_with_threshold = (
        (total_rod_length - scrap_below_threshold) * 100 / total_rod_length
        if total_rod_length > 0
        else 0
    )

    return {
        "total_rod_length": total_rod_length,
        "used_length": used_length,
        "loss": loss,
        "yield_rate": yield_rate,
        "yield_rate_with_threshold": yield_rate_with_threshold,
        "scrap_below_threshold": scrap_below_threshold,
        "scrap_above_threshold": scrap_above_threshold,
    }


def execute_optimizer(available_rods, required_cuts, diameter, time_limit, uploaded_file, input_method, scrap_threshold=400, reuse_rods=None):
    start_time = time.perf_counter()

    # 必要な切り出し長さを変数として定義
    l = [int(s) for s in required_cuts.keys()]
    q = [int(s) for s in required_cuts.values()]

    # 再利用端材の処理
    reuse_pattern_counts = []
    all_combinations = []

    if reuse_rods:
        # 各再利用材料からの切断パターンを生成
        for rod_length, rod_count in reuse_rods.items():
            rod_combinations = generate_all_combinations([rod_length], required_cuts)
            pattern_count = len(rod_combinations)
            reuse_pattern_counts.append({
                'rod_length': rod_length,
                'count': rod_count,
                'pattern_count': pattern_count
            })
            all_combinations.extend(rod_combinations)

    # 全組み合わせを計算（通常の材料）
    normal_combinations = generate_all_combinations(available_rods, required_cuts)
    all_combinations.extend(normal_combinations)
    combinations_count = len(all_combinations)

    # 結果を格納する辞書
    result_data = {
        "combinations_count": combinations_count,
        "available_rods": available_rods,
        "diameter": diameter,
        "uploaded_file": uploaded_file,
        "input_method": input_method,
    }

    if not all_combinations:
        result_data["error"] = "有効な組み合わせが見つかりませんでした"
        return result_data
    else:
        # 最適化問題用に変数を定義
        a = []
        c = []
        for combo in all_combinations:
            a.append([combo['cuts'].count(i) for i in l])
            c.append(combo['loss'])

        # 最適な切り出しプランを計算
        optimal_solution, optimal_value = optimal_cutting_plan(c, a, q, time_limit, reuse_pattern_counts)

        end_time = time.perf_counter()
        processing_time = end_time - start_time

        if optimal_solution is not None:
            total_rod_length = 0
            used_length = 0
            used_list = []
            cutting_patterns = []

            # 再利用端材の使用本数を追跡
            used_reuse_rods = {}
            if reuse_rods:
                for rod_length in reuse_rods.keys():
                    used_reuse_rods[rod_length] = 0

            for i in range(len(all_combinations)):
                if int(optimal_solution[i]) > 0:
                    pattern = {
                        'rod_length': all_combinations[i]['rod_length'],
                        'cuts': all_combinations[i]['cuts'],
                        'loss': all_combinations[i]['rod_length'] - sum(all_combinations[i]['cuts']),
                        'num': int(optimal_solution[i])
                    }
                    cutting_patterns.append(pattern)
                    total_rod_length += all_combinations[i]['rod_length'] * int(optimal_solution[i])
                    used_length += sum(all_combinations[i]['cuts']) * int(optimal_solution[i])
                    used_list.extend(all_combinations[i]['cuts'] * int(optimal_solution[i]))

                    # 再利用端材の使用本数をカウント
                    if reuse_rods and all_combinations[i]['rod_length'] in reuse_rods:
                        used_reuse_rods[all_combinations[i]['rod_length']] += int(optimal_solution[i])

            # 残りの再利用端材を計算
            remaining_reuse_rods = {}
            if reuse_rods:
                for rod_length, total_count in reuse_rods.items():
                    used_count = used_reuse_rods.get(rod_length, 0)
                    remaining_count = total_count - used_count
                    if remaining_count > 0:
                        remaining_reuse_rods[rod_length] = remaining_count

            # 検証
            used_count = [used_list.count(i) for i in l]

            if used_count == q:
                # 切断パターンDataFrameを作成
                df_results = pd.DataFrame([
                    {
                        'id': i + 1,
                        'times': pattern['num'],
                        'loss': pattern['loss'],
                        'base': pattern['rod_length'],
                        'pattern': ','.join(map(str, pattern['cuts'])),
                    }
                    for i, pattern in enumerate(cutting_patterns)
                ])

                # 結果シート作成
                expanded_df, count_df, cutting_count_df, sheet_cutting_data_all, cutting_unique_values = create_result_sheets(
                    df_results, diameter, uploaded_file if input_method == "XLSXファイルアップロード" else None
                )

                # 履歴に追加（基本的な計算）
                loss = total_rod_length - used_length
                yield_rate = used_length * 100 / total_rod_length

                history_record = {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "diameter": diameter,
                    "yield_rate": yield_rate,
                    "total_length": total_rod_length,
                    "loss": loss,
                    "time": processing_time,
                    "time_limit": time_limit,
                    "required_cuts": required_cuts,
                    "cutting_patterns": cutting_patterns,
                    "combinations_count": combinations_count,
                }
                st.session_state.history.append(history_record)

                # 結果データを格納（閾値計算は表示時に行う）
                result_data.update(
                    {
                        "success": True,
                        "processing_time": processing_time,
                        "total_rod_length": total_rod_length,
                        "used_length": used_length,
                        "cutting_patterns": cutting_patterns,
                        "df_results": df_results,
                        "expanded_df": expanded_df,
                        "count_df": count_df,
                        "cutting_count_df": cutting_count_df,
                        "sheet_cutting_data_all": sheet_cutting_data_all,
                        "cutting_unique_values": cutting_unique_values,
                        "remaining_reuse_rods": remaining_reuse_rods,
                    }
                )
            else:
                result_data["success"] = False
                result_data["error"] = (
                    f"要求本数と切り出し個数が一致しません。要求: {q}, 切り出し: {used_count}"
                )
        else:
            result_data["success"] = False
            result_data["error"] = "最適解が見つかりませんでした"
    return result_data

def display_optimization_results(result, scrap_threshold, tab_key=""):
    """最適化結果を表示する関数"""
    if not result.get("success"):
        st.error(result.get("error", "最適化に失敗しました"))
        return

    cutting_patterns = result["cutting_patterns"]
    processing_time = result["processing_time"]
    combinations_count = result["combinations_count"]
    df_results = result["df_results"]
    diameter = result["diameter"]
    uploaded_file = result.get("uploaded_file")
    input_method = result.get("input_method")
    expanded_df = result["expanded_df"]
    count_df = result["count_df"]
    cutting_count_df = result.get("cutting_count_df")
    sheet_cutting_data_all = result.get("sheet_cutting_data_all")
    cutting_unique_values = result.get("cutting_unique_values")

    # 閾値に基づいて再計算
    recalc_result = recalculate_with_threshold(cutting_patterns, scrap_threshold)
    st.write(f"**切断パターン組み合わせ:** {combinations_count:,}")
    st.success("最適化が完了しました！")

    # サマリー表示
    col_summary1, col_summary2, col_summary3 = st.columns([1, 1, 1])
    with col_summary1:
        st.metric("処理時間", f"{processing_time:.4f} s")
        st.metric("総材料長", f"{recalc_result['total_rod_length']:,} mm")
    with col_summary2:
        st.metric("歩留り率(再利用なし)", f"{recalc_result['yield_rate']:.2f}%")
        st.metric("端材(再利用なし)", f"{recalc_result['loss']:,} mm")
    with col_summary3:
        st.metric(
            "歩留り率(再利用あり)",
            f"{recalc_result['yield_rate_with_threshold']:.2f}%",
            delta=f"{recalc_result['yield_rate_with_threshold'] - recalc_result['yield_rate']:.2f}%",
        )
        st.metric("端材(再利用あり)", f"{recalc_result['scrap_below_threshold']:,} mm")

    # 切断パターンの表示
    st.write("切断パターン")
    st.dataframe(df_results, use_container_width=True)

    # 再利用端材の表示と処理
    st.write("再利用端材")
    scrap_above_threshold = recalc_result["scrap_above_threshold"]

    # 残った再利用端材を新しい端材リストに追加
    remaining_reuse_rods = result.get("remaining_reuse_rods", {})
    combined_scrap_dict = {}

    # 新しい端材を辞書に追加
    for item in scrap_above_threshold:
        length = item["length"]
        count = item["count"]
        combined_scrap_dict[length] = combined_scrap_dict.get(length, 0) + count

    # 残った再利用端材を追加
    for length, count in remaining_reuse_rods.items():
        combined_scrap_dict[length] = combined_scrap_dict.get(length, 0) + count

    # DataFrameに変換
    df_reusable_scrap = pd.DataFrame(
        [
            {"端材の長さ (mm)": length, "本数": count}
            for length, count in sorted(combined_scrap_dict.items(), reverse=True)
        ]
    )

    st.dataframe(df_reusable_scrap)

    # ダウンロードボタン
    st.write("ダウンロード")
    col_download1, col_download2 = st.columns([1, 1])
    with col_download1:
        # 閾値以上の端材のCSVダウンロード（新しい端材+残った再利用端材）
        if not df_reusable_scrap.empty:
            csv_reusable = df_reusable_scrap.to_csv(index=False, encoding="utf-8-sig")

            st.download_button(
                label=f"再利用端材リスト\n(≥{scrap_threshold}mm)",
                data=csv_reusable,
                file_name=f"reusable_scrap_{diameter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                key=f"reuse_{tab_key}_{scrap_threshold}",
            )
        else:
            st.info(f"再利用可能な端材(≥{scrap_threshold}mm)はありません")

    with col_download2:
        # Excelダウンロード
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # 最適切断パターンを最初のシートに保存
            expanded_df.to_excel(writer, sheet_name="最適切断パターン", index=False)

            # 切断種類のカウント結果を2番目のシートに保存
            count_df.to_excel(writer, sheet_name="出力結果集計表", index=False)

            # 切断集計表のカウント結果を3番目のシートに保存(データがある場合)
            if cutting_count_df is not None:
                cutting_count_df.to_excel(
                    writer, sheet_name="切断指示集計表", index=False
                )

        # Excel関数を設定
        excel_buffer.seek(0)
        workbook = openpyxl.load_workbook(excel_buffer)

        # 出力結果集計表に合計の数式を設定
        if "出力結果集計表" in workbook.sheetnames:
            ws_output = workbook["出力結果集計表"]
            last_row = ws_output.max_row

            # 合計行の数式を設定(B列から最後の列まで)
            for col in range(2, ws_output.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws_output[f"{col_letter}{last_row}"] = (
                    f"=SUM({col_letter}2:{col_letter}{last_row-1})"
                )

        # 切断指示集計表に合計の数式を設定
        if cutting_count_df is not None and "切断指示集計表" in workbook.sheetnames:
            ws_cutting = workbook["切断指示集計表"]
            last_row_cutting = ws_cutting.max_row

            # 合計行の数式を設定(B列から最後の列まで)
            for col in range(2, ws_cutting.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws_cutting[f"{col_letter}{last_row_cutting}"] = (
                    f"=SUM({col_letter}2:{col_letter}{last_row_cutting-1})"
                )

            # サマリーシートを作成
            ws_summary = workbook.create_sheet("サマリー")
            ws_summary["A1"] = "径"
            ws_summary["B1"] = diameter

            ws_summary["A2"] = "歩留り率(再利用なし)"
            ws_summary["B2"] = recalc_result["yield_rate"]
            ws_summary["B2"].number_format = "0.00%"
            ws_summary["B2"].value = recalc_result["yield_rate"] / 100

            ws_summary["A3"] = "端材(再利用なし)[mm]"
            ws_summary["B3"] = recalc_result["loss"]
            ws_summary["B3"].number_format = "#,##0"

            ws_summary["A4"] = "端材閾値"
            ws_summary["B4"] = scrap_threshold

            ws_summary["A5"] = "歩留り率(再利用あり)"
            ws_summary["B5"] = recalc_result["yield_rate_with_threshold"]
            ws_summary["B5"].number_format = "0.00%"
            ws_summary["B5"].value = recalc_result["yield_rate_with_threshold"] / 100

            ws_summary["A6"] = "端材(再利用あり)[mm]"
            ws_summary["B6"] = recalc_result["scrap_below_threshold"]
            ws_summary["B6"].number_format = "#,##0"

            ws_summary["A7"] = "総材料長"
            ws_summary["B7"] = recalc_result["total_rod_length"]
            ws_summary["B7"].number_format = "#,##0"

            ws_summary["A8"] = "処理時間(s)"
            ws_summary["B8"] = processing_time
            ws_summary["B8"].number_format = "#,##0.00"

            # 切断指示集計表がある場合
            if cutting_unique_values:
                ws_summary["A10"] = "差分(出力結果 - 切断指示)"

                # 差分のヘッダーを設定
                ws_summary["A11"] = "長さ(mm)"
                ws_summary["B11"] = "出力結果"
                ws_summary["C11"] = "切断指示"
                ws_summary["D11"] = "差分"

                # 差分の数式を設定
                row_num = 12
                for i, value in enumerate(cutting_unique_values):
                    col_letter_cutting = openpyxl.utils.get_column_letter(
                        i + 2
                    )  # 切断指示集計表のB列から開始
                    col_letter_output = openpyxl.utils.get_column_letter(
                        i + 2
                    )  # 出力結果集計表のB列から開始

                    ws_summary[f"A{row_num}"] = str(value)
                    ws_summary[f"B{row_num}"] = (
                        f"=出力結果集計表!{col_letter_output}{last_row}"
                    )
                    ws_summary[f"C{row_num}"] = (
                        f"=切断指示集計表!{col_letter_cutting}{last_row_cutting}"
                    )
                    ws_summary[f"D{row_num}"] = f"=B{row_num}-C{row_num}"
                    row_num += 1

        # 修正されたExcelファイルを保存
        new_buffer = io.BytesIO()
        workbook.save(new_buffer)
        workbook.close()
        new_buffer.seek(0)

        st.download_button(
            label="最適化結果シート",
            data=new_buffer.getvalue(),
            file_name=f"result_sheet_{diameter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"excel_{tab_key}_{scrap_threshold}",
        )

def main():
    st.set_page_config(page_title="鉄筋切断最適化アプリ", layout="wide")

    st.title("🔧 鉄筋切断最適化アプリ")
    st.write("鉄筋の切断パターンを最適化し、廃材を最小化します。")

    # サイドバーで実行履歴を表示
    with st.sidebar:
        st.header("📊 実行履歴")

        # セッション状態の初期化
        if 'history' not in st.session_state:
            st.session_state.history = []

        if st.session_state.history:
            for i, record in enumerate(reversed(st.session_state.history[-10:])):  # 最新10件のみ表示
                with st.expander(f"{record['timestamp']} - {record['diameter']}"):
                    st.write(f"**径:** {record['diameter']}")
                    st.write(f"**歩留り率:** {record['yield_rate']:.2f}%")
                    st.write(f"**総材料長:** {record['total_length']:,} mm")
                    st.write(f"**端材:** {record['loss']:,} mm")
                    st.write(f"**処理時間:** {record['time']:.4f} s")
                    st.write(f"**制限時間:** {record.get('time_limit', 120)} s")
                    st.write(f"**組み合わせ数:** {record['combinations_count']:,}")

                    # 必要な切り出しを表示
                    st.write("**必要な切り出し:**")
                    for length, count in sorted(record['required_cuts'].items(), reverse=True):
                        st.write(f"• {length}mm × {count}本")

                    # 切断パターンを見やすく表示
                    display_cutting_patterns(record['cutting_patterns'])
        else:
            st.write("実行履歴がありません")

        # 履歴のクリアボタン
        if st.button("履歴をクリア"):
            st.session_state.history = []
            st.rerun()

    # メインコンテンツ
    col1, col2 = st.columns([1, 1])

    with col1:
        st.header("⚙️ 設定")

        # 径の選択
        diameter = st.selectbox(
            "鉄筋径を選択してください:",
            options=list(BASE_PATTERNS.keys())
        )

        # 利用可能な棒の長さを表示
        st.write(f"**{diameter}の利用可能な棒の長さ:** {BASE_PATTERNS[diameter]}")

        # 実行時間の上限設定
        time_limit = st.number_input(
            "最適化の制限時間 (10~3600 秒)",
            min_value=10,
            max_value=3600,
            value=120,
            step=10,
            help="最適化計算の制限時間を設定します。大きな問題では時間を長く設定することを推奨します。"
        )
        st.write(f"現在の設定: {time_limit}秒")

        # 入力方法の選択
        input_method = st.radio(
            "入力方法を選択してください:",
            ("XLSXファイルアップロード", "手入力")
        )

        required_cuts = {}
        uploaded_file = None
        reuse_rods = {}

        # 再利用端材のCSVアップロード
        st.subheader("再利用端材 (オプション)")
        scrap_csv_file = st.file_uploader(
            "再利用端材のCSVファイル (オプション)",
            type=['csv'],
            help="端材の長さ (mm),本数 の形式のCSVファイルをアップロードしてください",
            key="scrap_csv_uploader"
        )

        if scrap_csv_file is not None:
            with st.spinner("端材CSVを解析中..."):
                scrap_data = read_scrap_data_from_csv(scrap_csv_file)

            if scrap_data:
                reuse_rods = scrap_data
                st.success(f"再利用端材のCSVファイルを正常に読み込みました。")

                # データの詳細表示
                st.write("**読み込まれた再利用端材:**")
                df_scrap_preview = pd.DataFrame([
                    {'端材の長さ (mm)': length, '本数': count}
                    for length, count in sorted(scrap_data.items(), reverse=True)
                ])
                st.dataframe(df_scrap_preview, use_container_width=True)

                # 統計情報
                total_scrap_pieces = sum(scrap_data.values())
                scrap_types = len(scrap_data)
                st.write(f"**統計情報:** 端材種類数: {scrap_types}種類, 総端材本数: {total_scrap_pieces:,}本")
            else:
                st.error("再利用端材のCSVファイルの読み込みに失敗しました。")

        if input_method == "XLSXファイルアップロード":
            st.subheader("XLSXファイルアップロード")

            uploaded_file = st.file_uploader(
                "XLSXファイルを選択してください",
                type=['xlsx'],
                help="切断集計表のXLSXファイルをアップロードしてください"
            )

            if uploaded_file is not None:
                with st.spinner("ファイルを解析中..."):
                    cutting_data = read_cutting_data_from_xlsx(uploaded_file, diameter)

                if cutting_data:
                    required_cuts = cutting_data
                    st.success(f"XLSXファイルを正常に読み込みました。{diameter}のデータが見つかりました。")

                    # データの詳細表示
                    st.write("**読み込まれたデータ:**")
                    df_preview = pd.DataFrame([
                        {'長さ (mm)': length, '本数': count}
                        for length, count in sorted(cutting_data.items(), reverse=True)
                    ])
                    st.dataframe(df_preview, use_container_width=True)

                    # 統計情報
                    total_pieces = sum(cutting_data.values())
                    cutting_types = len(cutting_data)
                    st.write(f"**統計情報:** 切断種類数: {cutting_types}種類, 総切断本数: {total_pieces:,}本")
                else:
                    st.error(f"XLSXファイルに{diameter}のデータが見つかりませんでした。")

        else:  # 手入力
            st.subheader("必要な切り出し長さと本数")

            # 動的入力フィールド
            if 'input_rows' not in st.session_state:
                st.session_state.input_rows = 5

            for i in range(st.session_state.input_rows):
                col_length, col_count = st.columns([1, 1])

                with col_length:
                    length = st.number_input(
                        f"長さ {i+1} (mm)",
                        min_value=0,
                        value=0,
                        key=f"length_{i}"
                    )

                with col_count:
                    count = st.number_input(
                        f"本数 {i+1}",
                        min_value=0,
                        value=0,
                        key=f"count_{i}"
                    )

                if length > 0 and count > 0:
                    required_cuts[length] = count

            col_add, col_remove = st.columns([1, 1])
            with col_add:
                if st.button("行を追加"):
                    st.session_state.input_rows += 1
                    st.rerun()

            with col_remove:
                if st.button("行を削除") and st.session_state.input_rows > 1:
                    st.session_state.input_rows -= 1
                    st.rerun()

            # 手入力データのプレビュー
            if required_cuts:
                st.write("**入力されたデータ:**")
                df_preview = pd.DataFrame([
                    {'長さ (mm)': length, '本数': count}
                    for length, count in sorted(required_cuts.items(), reverse=True)
                ])
                st.dataframe(df_preview, use_container_width=True)

    with col2:
        st.header("🎯 最適化結果")

        scrap_threshold = st.number_input(
            "端材閾値 (mm)",
            min_value=0,
            max_value=2000,
            value=400,
            step=50,
            help="この閾値未満の端材のみを廃材として歩留り率を計算します。閾値以上の端材は再利用可能として扱われます。"
        )
        st.write(f"現在の設定: {scrap_threshold}mm未満を廃材として扱う")

        # session_stateの初期化
        if 'optimization_results' not in st.session_state:
            st.session_state.optimization_results = {}

        if st.button("最適化を実行", type="primary") and required_cuts:
            with st.spinner("最適化を実行中..."):
                # 利用可能な棒の長さを取得
                available_rods = BASE_PATTERNS[diameter].copy()
                # 結果をsession_stateに保存
                st.session_state.optimization_results = {}

                # 複数材料での最適化
                result_all = execute_optimizer(available_rods, required_cuts, diameter, time_limit, uploaded_file, input_method, scrap_threshold, reuse_rods if reuse_rods else None)
                st.session_state.optimization_results['all'] = result_all

                # 単一材料での最適化
                for rod in available_rods:
                    result_single = execute_optimizer([rod], required_cuts, diameter, time_limit, uploaded_file, input_method, scrap_threshold, reuse_rods if reuse_rods else None)
                    st.session_state.optimization_results[f'{rod}mm'] = result_single

        # 結果がある場合は表示
        if st.session_state.optimization_results:
            available_rods = BASE_PATTERNS[diameter].copy()
            tab_names = ["全種類"] + [f"{i} mm" for i in available_rods]
            tabs = st.tabs(tab_names)

            # 複数材料での結果表示
            with tabs[0]:
                if 'all' in st.session_state.optimization_results:
                    display_optimization_results(st.session_state.optimization_results['all'], scrap_threshold, "all")

            # 単一材料での結果表示
            for i, rod in enumerate(available_rods):
                with tabs[i+1]:
                    if f'{rod}mm' in st.session_state.optimization_results:
                        display_optimization_results(st.session_state.optimization_results[f'{rod}mm'], scrap_threshold, f"{rod}mm")

        elif not required_cuts:
            st.info("切断指示を入力してください。")
        else:
            st.info("最適化を実行してください。")

    # フッター
    st.markdown("---")
    st.markdown("💡 **使用方法:**")
    st.markdown("1. 鉄筋の径を選択")
    st.markdown("2. XLSXファイルアップロードまたは手入力で切断指示を入力")
    st.markdown("3. 最適化を実行")
    st.markdown("4. 出力結果をExcelファイルでダウンロード")

if __name__ == "__main__":
    main()
