import streamlit as st
import pandas as pd
import time
import pulp
from datetime import datetime
import openpyxl
from collections import defaultdict
import io

BASE_PATTERNS = {
    'D10': [4000, 4500, 5500, 6000],
    'D13': [3000, 4000, 4500, 5500, 6000, 7500],
    'D16': [4000, 4500, 5500, 6000, 7000],
    'D19': [3500, 4000, 4500, 5500, 6000],
    'D22': [4000, 4500, 5500, 6000]
}

# 全ての径のリスト
DIAMETERS = ["D10", "D13", "D16", "D19", "D22"]

# 全ての径で選択可能な長さのリスト
ALL_AVAILABLE_LENGTHS = sorted(
    list(set(length for lengths in BASE_PATTERNS.values() for length in lengths)), reverse=True
)

def find_cell_position(worksheet, search_text):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value and search_text in str(cell.value):
                return cell.row, cell.column
    return None

def get_diameter_column_index(worksheet, base_row, base_col, target_diameter):
    col = base_col + 1
    max_col = worksheet.max_column
    
    while col <= max_col:
        cell_value = worksheet.cell(row=base_row, column=col).value
        if cell_value and str(cell_value).strip() == target_diameter:
            return col
        col += 1
    
    return None

def extract_cutting_data_from_sheet(worksheet, target_diameter):
    base_position = find_cell_position(worksheet, "鉄筋径")
    if not base_position:
        return {}
    
    base_row, base_col = base_position
    diameter_col = get_diameter_column_index(worksheet, base_row, base_col, target_diameter)
    if not diameter_col:
        return {}
    
    cutting_data = {}
    row = base_row + 1
    max_row = worksheet.max_row
    
    while row <= max_row:
        length_cell = worksheet.cell(row=row, column=base_col+1)
        count_cell = worksheet.cell(row=row, column=diameter_col)
        
        if (length_cell.value is not None and count_cell.value is not None and
            isinstance(length_cell.value, (int, float)) and 
            isinstance(count_cell.value, (int, float)) and
            count_cell.value > 0):
            
            length = int(length_cell.value)
            count = int(count_cell.value)
            cutting_data[length] = count
        
        if length_cell.value is None or length_cell.value == '':
            empty_count = 0
            temp_row = row
            while temp_row <= min(row + 2, max_row):
                if worksheet.cell(row=temp_row, column=base_col).value is None:
                    empty_count += 1
                temp_row += 1
            if empty_count >= 3:
                break
        
        row += 1
    
    return cutting_data

def create_result_sheets(df_results, target_diameter, uploaded_file=None):
    max_items = 0
    for pattern in df_results['pattern']:
        if pd.notna(pattern):
            items = [item.strip().replace('"', '') for item in str(pattern).split(',')]
            max_items = max(max_items, len(items))

    df = df_results.copy()
    
    for i in range(1, max_items + 1):
        df[f'item_{i}'] = ''

    for index, row in df.iterrows():
        pattern = row['pattern']
        if pd.notna(pattern):
            items = [item.strip().replace('"', '') for item in str(pattern).split(',')]
            for i, item in enumerate(items):
                df.at[index, f'item_{i + 1}'] = item

    df = df.drop('pattern', axis=1)

    unique_values = set()
    for i in range(1, max_items + 1):
        col_name = f'item_{i}'
        for value in df[col_name]:
            if value and value.strip():
                unique_values.add(value.strip())

    unique_values = sorted(list(unique_values), key=lambda x: int(x) if x.isdigit() else float('inf'), reverse=True)

    count_data = []
    for index, row in df.iterrows():
        id_value = row['id']
        times_value = int(row['times']) if pd.notna(row['times']) else 0
        
        count_row = {'id': id_value}
        for value in unique_values:
            count_row[value] = 0
        
        for i in range(1, max_items + 1):
            col_name = f'item_{i}'
            item_value = row[col_name]
            if item_value and item_value.strip():
                item_value = item_value.strip()
                if item_value in count_row:
                    count_row[item_value] += times_value
        
        count_data.append(count_row)

    count_df = pd.DataFrame(count_data)

    cutting_total_row = {'id': '合計'}
    for value in unique_values:
        cutting_total_row[str(value)] = None
    count_df = pd.concat([count_df, pd.DataFrame([cutting_total_row])], ignore_index=True)

    cutting_count_df = None
    sheet_cutting_data_all = None
    cutting_unique_values = None
    
    if uploaded_file:
        try:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            
            sheet_cutting_data_all = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                sheet_cutting_data = extract_cutting_data_from_sheet(worksheet, target_diameter)
                
                if sheet_cutting_data:
                    sheet_cutting_data_all[sheet_name] = sheet_cutting_data
            
            workbook.close()
            
            if sheet_cutting_data_all:
                cutting_unique_values = set()
                for sheet_data in sheet_cutting_data_all.values():
                    cutting_unique_values.update(sheet_data.keys())
                
                cutting_unique_values = sorted(list(cutting_unique_values), 
                                             key=lambda x: int(x) if str(x).isdigit() else float('inf'), reverse=True)
                
                cutting_count_data = []
                for sheet_name, sheet_data in sheet_cutting_data_all.items():
                    count_row = {'シート名': sheet_name}
                    for value in cutting_unique_values:
                        count_row[str(value)] = sheet_data.get(value, 0)
                    
                    cutting_count_data.append(count_row)
                
                cutting_count_df = pd.DataFrame(cutting_count_data)
                
                project_total_row = {'シート名': '合計'}
                for value in cutting_unique_values:
                    project_total_row[str(value)] = None
                cutting_count_df = pd.concat([cutting_count_df, pd.DataFrame([project_total_row])], ignore_index=True)
        
        except Exception as e:
            st.warning(f"切断集計表の処理中にエラーが発生しました: {e}")

    return df, count_df, cutting_count_df, sheet_cutting_data_all, cutting_unique_values

def read_cutting_data_from_xlsx(uploaded_file, target_diameter):
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        total_cutting_data = defaultdict(int)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_cutting_data = extract_cutting_data_from_sheet(worksheet, target_diameter)

            for length, count in sheet_cutting_data.items():
                total_cutting_data[length] += count

        workbook.close()
        return dict(total_cutting_data)

    except Exception as e:
        st.error(f"XLSXファイルの読み込み中にエラーが発生しました: {e}")
        return {}

def read_scrap_data_from_csv(uploaded_file):
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()

        # 3列形式を想定：鉄筋径、長さ、本数
        if len(df.columns) < 3:
            st.error("CSVファイルには3列(鉄筋径、長さ、本数)が必要です")
            return {}

        scrap_data_by_diameter = {}
        for _, row in df.iterrows():
            diameter = str(row.iloc[0]).strip()
            length = row.iloc[1]
            count = row.iloc[2]

            if pd.notna(length) and pd.notna(count):
                try:
                    length_int = int(float(length))
                    count_int = int(float(count))
                    if length_int > 0 and count_int > 0:
                        if diameter not in scrap_data_by_diameter:
                            scrap_data_by_diameter[diameter] = {}
                        scrap_data_by_diameter[diameter][length_int] = \
                            scrap_data_by_diameter[diameter].get(length_int, 0) + count_int
                except (ValueError, TypeError):
                    continue

        return scrap_data_by_diameter

    except Exception as e:
        st.error(f"CSVファイルの読み込み中にエラーが発生しました: {e}")
        return {}

def validate_material_selection(cutting_data, selected_materials):
    """
    切断指示書の内容と選択された材料長をバリデーション
    
    Parameters:
    - cutting_data: 径ごとの切断指示データ {径: {長さ: 本数}}
    - selected_materials: 径ごとの選択された材料長 {径: [長さのリスト]}
    
    Returns:
    - is_valid: バリデーション結果 (True/False)
    - error_messages: エラーメッセージのリスト
    """
    error_messages = []
    
    # 切断指示書に含まれる径をチェック
    for diameter, cuts in cutting_data.items():
        if not cuts:
            continue
        
        # 1. この径の材料長が一つも選択されていないかチェック
        selected_lengths = selected_materials.get(diameter, [])
        if not selected_lengths:
            cut_lengths = sorted(cuts.keys(), reverse=True)
            error_messages.append(
                f"❌ **{diameter}**: 切断指示書に {len(cuts)}種類の切断長さ（最大{max(cut_lengths)}mm）がありますが、材料長が選択されていません"
            )
            continue
        
        # 2. 最大切断長さをチェック
        max_cut_length = max(cuts.keys())
        max_selected_length = max(selected_lengths)
        
        if max_cut_length > max_selected_length:
            # 対応可能な最小の材料長を提案
            available_in_base = [length for length in BASE_PATTERNS.get(diameter, []) if length >= max_cut_length]
            suggestion = f"推奨: {min(available_in_base)}mm以上" if available_in_base else "対応可能な材料長がありません"
            
            error_messages.append(
                f"❌ **{diameter}**: 最大切断長さ {max_cut_length}mm に対して、選択された材料長の最大値 {max_selected_length}mm では対応できません（{suggestion}）"
            )
    
    is_valid = len(error_messages) == 0
    return is_valid, error_messages

def dfs(index, current_combination, current_sum, remaining_counts, sorted_numbers, max_sum, all_combinations):
    if 0 < current_sum <= max_sum:
        sorted_combo = tuple(sorted(current_combination, reverse=True))
        all_combinations.add((sorted_combo, current_sum))
    
    if current_sum > max_sum or index >= len(sorted_numbers):
        return
    
    current_number = sorted_numbers[index]
    max_count = remaining_counts[current_number]
    
    for count in range(max_count + 1):
        new_sum = current_sum + current_number * count
        
        if new_sum > max_sum:
            break
        
        new_combination = current_combination + [current_number] * count
        new_remaining = remaining_counts.copy()
        new_remaining[current_number] -= count
        
        dfs(index + 1, new_combination, new_sum, new_remaining, sorted_numbers, max_sum, all_combinations)

def find_combinations_dfs(numbers_dict, max_sum=7500):
    sorted_numbers = sorted(numbers_dict.keys(), reverse=True)
    all_combinations = set()
    
    dfs(0, [], 0, numbers_dict.copy(), sorted_numbers, max_sum, all_combinations)
    
    return [(list(combo), sum_val) for combo, sum_val in all_combinations]

def generate_all_combinations(available_rods, required_cuts):
    combinations = find_combinations_dfs(required_cuts, max(available_rods))
    combinations.sort(key=lambda x: x[1])

    available_rods.sort()
    i = 0

    all_combinations = []
    for combo, total_cut_length in combinations:
        while available_rods[i] < total_cut_length:
            i += 1
        loss = (available_rods[i] - total_cut_length) / available_rods[i]
        all_combinations.append({
            'rod_length': available_rods[i],
            'cuts': tuple(combo),
            'loss': loss,
        })
    
    all_combinations.sort(key=lambda x: x['cuts'], reverse=True)
    all_combinations.sort(key=lambda x: x['rod_length'], reverse=False)
    return all_combinations

def optimal_cutting_plan(c, a, q, time_limit=120, reuse_pattern_counts=None):
    n = len(a)
    m = len(q)

    prob = pulp.LpProblem("CuttingStockProblem", pulp.LpMinimize)
    x = [pulp.LpVariable(f"y{j+1}", lowBound=0, cat='Integer') for j in range(n)]

    if reuse_pattern_counts:
        pattern_index = 0
        for reuse_info in reuse_pattern_counts:
            pattern_count = reuse_info['pattern_count']
            rod_count = reuse_info['count']

            if pattern_count > 0:
                reuse_constraint = pulp.lpSum(x[j] for j in range(pattern_index, pattern_index + pattern_count))
                prob += reuse_constraint <= rod_count, f"Reuse_constraint_{pattern_index}"

            pattern_index += pattern_count

    objective = pulp.lpSum(c[j] * x[j] for j in range(n))
    prob += objective, "Total_Loss"

    for i in range(m):
        production_constraint = pulp.lpSum(a[j][i] * x[j] for j in range(n))
        prob += production_constraint == q[i], f"Demand_constraint_{i+1}"

    prob.solve(pulp.PULP_CBC_CMD(timeLimit=time_limit, msg=False))

    if prob.status == pulp.LpStatusOptimal:
        optimal_solution = [var.varValue for var in x]
        optimal_value = pulp.value(prob.objective)
        return optimal_solution, optimal_value
    else:
        return None, None

def recalculate_with_threshold(cutting_patterns, scrap_threshold):
    total_rod_length = 0
    used_length = 0

    for pattern in cutting_patterns:
        total_rod_length += pattern["rod_length"] * pattern["num"]
        used_length += sum(pattern["cuts"]) * pattern["num"]

    loss = total_rod_length - used_length
    yield_rate = used_length * 100 / total_rod_length if total_rod_length > 0 else 0

    scrap_below_threshold = 0
    scrap_above_threshold = []

    for pattern in cutting_patterns:
        pattern_loss = pattern["loss"]
        pattern_num = pattern["num"]
        if pattern_loss < scrap_threshold:
            scrap_below_threshold += pattern_loss * pattern_num
        else:
            scrap_above_threshold.append(
                {
                    "length": pattern_loss,
                    "count": pattern_num,
                    "rod_length": pattern["rod_length"],
                }
            )

    yield_rate_with_threshold = (
        (total_rod_length - scrap_below_threshold) * 100 / total_rod_length
        if total_rod_length > 0
        else 0
    )

    return {
        "total_rod_length": total_rod_length,
        "used_length": used_length,
        "loss": loss,
        "yield_rate": yield_rate,
        "yield_rate_with_threshold": yield_rate_with_threshold,
        "scrap_below_threshold": scrap_below_threshold,
        "scrap_above_threshold": scrap_above_threshold,
    }

def consolidate_results_by_sheet(optimization_results):
    """
    複数径の最適化結果をシート単位で統合する
    在庫の取り出しと保管の操作も含める（径ごとに個別の行）
    シート順序は元のXLSXファイルの順番を維持
    
    Parameters:
    - optimization_results: 径ごとの最適化結果の辞書
    
    Returns:
    - consolidated_df: シート×径で統合されたDataFrame（在庫操作含む）
    - sheet_order: シートの処理順序リスト（元の順序）
    - sheet_color_map: シート名と色のマッピング
    """
    all_patterns = []

    # 各径の結果を収集
    for diameter, result in optimization_results.items():
        if not result.get("success"):
            continue

        df_results = result.get("df_results")
        if df_results is None or df_results.empty:
            continue

        # 径情報を追加
        df_copy = df_results.copy()
        df_copy['径'] = diameter
        all_patterns.append(df_copy)

    if not all_patterns:
        return pd.DataFrame(), [], {}

    # 全パターンを統合
    combined_df = pd.concat(all_patterns, ignore_index=True)

    # シート名の出現順序を取得（ソートしない、最初の径の順序を使用）
    sheet_order = []
    for sheet_name in combined_df['シート名']:
        if sheet_name not in sheet_order and sheet_name != '' and pd.notna(sheet_name):
            sheet_order.append(sheet_name)

    # シート名ごとに色を割り当て
    colors = ['#e6f2ff', '#ffe6f2', '#f2ffe6', '#fff2e6', '#f2e6ff', 
              '#e6fff2', '#ffe6e6', '#f2f2ff', '#fff9e6', '#ffe6ff']
    sheet_color_map = {}
    for idx, sheet_name in enumerate(sheet_order):
        if sheet_name == '未割り当て':
            sheet_color_map[sheet_name] = '#ffcccc'
        else:
            sheet_color_map[sheet_name] = colors[idx % len(colors)]

    # シート順序のマッピングを作成
    sheet_order_map = {name: idx for idx, name in enumerate(sheet_order)}
    combined_df['シート順序'] = combined_df['シート名'].map(sheet_order_map)

    # 径の順序マッピング
    diameter_order = {'D10': 1, 'D13': 2, 'D16': 3, 'D19': 4, 'D22': 5}
    combined_df['径順序'] = combined_df['径'].map(diameter_order)

    # '操作'列が存在するか確認
    has_operation = '操作' in combined_df.columns

    if not has_operation:
        # 操作列がない場合は全て切断として扱う
        cutting_df = combined_df.copy()
        grouped_df = cutting_df.groupby(['シート名', 'シート順序', '径', '径順序', 'base', 'pattern', 'loss'], as_index=False).size()
        grouped_df.rename(columns={'size': '本数'}, inplace=True)
        grouped_df = grouped_df.sort_values(by=['シート順序', '径順序'])
        grouped_df['統合切断順序'] = range(1, len(grouped_df) + 1)
        grouped_df['操作'] = '切断'
        grouped_df = grouped_df.drop(['径順序', 'シート順序'], axis=1)
        return grouped_df, sheet_order, sheet_color_map

    # 操作別に処理
    result_rows = []
    consolidated_order = 1

    for sheet_name in sheet_order:
        sheet_df = combined_df[combined_df['シート名'] == sheet_name]

        if sheet_df.empty:
            continue

        # 1. 在庫取り出し操作を径ごとに追加
        withdrawal_df = sheet_df[sheet_df['操作'] == '取出']
        if not withdrawal_df.empty:
            # 径でソートして追加
            for diameter in sorted(withdrawal_df['径'].unique(), key=lambda x: diameter_order.get(x, 999)):
                diameter_withdrawal = withdrawal_df[withdrawal_df['径'] == diameter]
                for _, row in diameter_withdrawal.iterrows():
                    result_rows.append({
                        '統合切断順序': consolidated_order,
                        'シート名': sheet_name,
                        '径': diameter,
                        '操作': '取出',
                        '本数': None,
                        'base': None,
                        'pattern': row['pattern'],
                        'loss': 0
                    })
                    consolidated_order += 1

        # 2. 切断パターンを追加
        cutting_df = sheet_df[sheet_df['操作'] == '切断']
        if not cutting_df.empty:
            # 径ごとにグループ化
            for diameter in sorted(cutting_df['径'].unique(), key=lambda x: diameter_order.get(x, 999)):
                diameter_df = cutting_df[cutting_df['径'] == diameter]

                # パターンごとにグループ化して本数をカウント
                grouped = diameter_df.groupby(['base', 'pattern', 'loss'], as_index=False).size()
                grouped.rename(columns={'size': '本数'}, inplace=True)

                for _, row in grouped.iterrows():
                    result_rows.append({
                        '統合切断順序': consolidated_order,
                        'シート名': sheet_name,
                        '径': diameter,
                        '操作': '切断',
                        '本数': row['本数'],
                        'base': row['base'],
                        'pattern': row['pattern'],
                        'loss': row['loss']
                    })
                    consolidated_order += 1

        # 3. 在庫保管操作を径ごとに追加
        storage_df = sheet_df[sheet_df['操作'] == '保管']
        if not storage_df.empty:
            # 径でソートして追加
            for diameter in sorted(storage_df['径'].unique(), key=lambda x: diameter_order.get(x, 999)):
                diameter_storage = storage_df[storage_df['径'] == diameter]
                for _, row in diameter_storage.iterrows():
                    result_rows.append({
                        '統合切断順序': consolidated_order,
                        'シート名': sheet_name,
                        '径': diameter,
                        '操作': '保管',
                        '本数': None,
                        'base': None,
                        'pattern': row['pattern'],
                        'loss': 0
                    })
                    consolidated_order += 1

    # 未割り当てパターンを追加
    unassigned_df = combined_df[combined_df['シート名'] == '未割り当て']
    if not unassigned_df.empty:
        cutting_df = unassigned_df[unassigned_df['操作'] == '切断']

        for diameter in sorted(cutting_df['径'].unique(), key=lambda x: diameter_order.get(x, 999)):
            diameter_df = cutting_df[cutting_df['径'] == diameter]

            grouped = diameter_df.groupby(['base', 'pattern', 'loss'], as_index=False).size()
            grouped.rename(columns={'size': '本数'}, inplace=True)

            for _, row in grouped.iterrows():
                result_rows.append({
                    '統合切断順序': consolidated_order,
                    'シート名': '未割り当て',
                    '径': diameter,
                    '操作': '切断',
                    '本数': row['本数'],
                    'base': row['base'],
                    'pattern': row['pattern'],
                    'loss': row['loss']
                })
                consolidated_order += 1

    consolidated_result_df = pd.DataFrame(result_rows)

    # 本数とbaseを整数型に変換（Noneは保持）
    consolidated_result_df['本数'] = consolidated_result_df['本数'].astype('Int64')
    consolidated_result_df['base'] = consolidated_result_df['base'].astype('Int64')

    return consolidated_result_df, sheet_order, sheet_color_map

def assign_patterns_to_sheets(cutting_patterns, sheet_cutting_data_all, initial_inventory=None):
    """
    切断パターンを各シートに割り当て、実際の施工フローに沿った切断順序を構成する
    在庫の取り出しと保管の操作も記録する
    
    Parameters:
    - cutting_patterns: 最適化で得られた切断パターンのリスト
    - sheet_cutting_data_all: 各シートの切断指示データ（OrderedDict等で順序保持）
    - initial_inventory: 初期在庫（再利用端材）
    
    Returns:
    - assigned_patterns: シート名、切断順序、在庫操作が追加されたパターンのリスト
    """
    if not sheet_cutting_data_all:
        return cutting_patterns
    
    # 各シートの残り切断指示を管理
    remaining_cuts = {}
    for sheet_name, sheet_data in sheet_cutting_data_all.items():
        remaining_cuts[sheet_name] = sheet_data.copy()
    
    # 在庫管理（切断で生成されたがまだ使用されていない材料）
    inventory = defaultdict(int)
    
    # 初期在庫があれば設定
    if initial_inventory:
        for length, count in initial_inventory.items():
            inventory[length] = count
    
    # 切断パターンを拡張（num回分を個別のパターンに）
    available_patterns = []
    for pattern in cutting_patterns:
        for _ in range(pattern['num']):
            available_patterns.append({
                'rod_length': pattern['rod_length'],
                'cuts': pattern['cuts'],
                'loss': pattern['loss']
            })
    
    # 割り当て結果
    assigned_patterns = []
    cutting_order = 1
    
    # 最初に初期在庫を記録（初期在庫がある場合）
    if initial_inventory and any(initial_inventory.values()):
        inventory_items = {length: count for length, count in initial_inventory.items() if count > 0}
        if inventory_items:
            assigned_patterns.append({
                'cutting_order': cutting_order,
                'sheet_name': '初期在庫',
                'operation': '保管',
                'inventory_items': inventory_items,
                'rod_length': None,
                'cuts': tuple(),
                'loss': 0
            })
            cutting_order += 1
    
    # シートごとに順番に処理（元の順序を維持）
    for sheet_name in sheet_cutting_data_all.keys():
        # ステップ1: このシート開始前の在庫取り出し操作を記録
        withdrawn_items = {}
        for length in list(remaining_cuts[sheet_name].keys()):
            need = remaining_cuts[sheet_name].get(length, 0)
            available = inventory.get(length, 0)
            use = min(need, available)
            if use > 0:
                withdrawn_items[length] = use
                inventory[length] -= use
                remaining_cuts[sheet_name][length] -= use
        
        # 在庫取り出し操作を記録（材料があった場合のみ）
        if withdrawn_items:
            assigned_patterns.append({
                'cutting_order': cutting_order,
                'sheet_name': sheet_name,
                'operation': '取出',
                'inventory_items': withdrawn_items,
                'rod_length': None,
                'cuts': tuple(),
                'loss': 0
            })
            cutting_order += 1
        
        # ステップ2: このシートの切断パターンを実行
        while any(remaining_cuts[sheet_name].get(length, 0) > 0 for length in remaining_cuts[sheet_name]):
            # このシートに最も貢献する切断パターンを選択
            best_idx = None
            best_score = -1
            
            for idx, pattern in enumerate(available_patterns):
                # このパターンがこのシートの不足材料をいくつ含むか計算
                score = sum(1 for cut in pattern['cuts'] 
                           if remaining_cuts[sheet_name].get(cut, 0) > 0)
                
                if score > best_score:
                    best_score = score
                    best_idx = idx
            
            if best_idx is None or best_score == 0:
                # このシートの切断指示を満たせるパターンがない場合
                break
            
            # 切断を実行して結果に追加
            pattern = available_patterns.pop(best_idx)
            assigned_patterns.append({
                'cutting_order': cutting_order,
                'sheet_name': sheet_name,
                'operation': '切断',
                'rod_length': pattern['rod_length'],
                'cuts': pattern['cuts'],
                'loss': pattern['loss']
            })
            
            # 切断で得られた材料を一時的に保持
            temp_inventory = defaultdict(int)
            for cut in pattern['cuts']:
                temp_inventory[cut] += 1
            
            # このシートで使用する分を差し引き
            for length in list(temp_inventory.keys()):
                need = remaining_cuts[sheet_name].get(length, 0)
                use = min(need, temp_inventory[length])
                if use > 0:
                    temp_inventory[length] -= use
                    remaining_cuts[sheet_name][length] -= use
            
            # 残った材料を在庫に追加
            for length, count in temp_inventory.items():
                if count > 0:
                    inventory[length] += count
            
            cutting_order += 1
        
        # ステップ3: このシート終了後の保管操作を記録（在庫がある場合のみ）
        if any(inventory.get(length, 0) > 0 for length in inventory):
            stored_items = {length: count for length, count in inventory.items() if count > 0}
            assigned_patterns.append({
                'cutting_order': cutting_order,
                'sheet_name': sheet_name,
                'operation': '保管',
                'inventory_items': stored_items,
                'rod_length': None,
                'cuts': tuple(),
                'loss': 0
            })
            cutting_order += 1
    
    # 残った切断パターンがあれば追加（未割り当て）
    for pattern in available_patterns:
        assigned_patterns.append({
            'cutting_order': cutting_order,
            'sheet_name': '未割り当て',
            'operation': '切断',
            'rod_length': pattern['rod_length'],
            'cuts': pattern['cuts'],
            'loss': pattern['loss']
        })
        cutting_order += 1
    
    # 最終的な在庫を返す値に含める
    final_inventory = {length: count for length, count in inventory.items() if count > 0}
    
    return assigned_patterns, final_inventory

def execute_optimizer(available_rods, required_cuts, diameter, time_limit, uploaded_file, input_method, scrap_threshold=400, reuse_rods=None):
    start_time = time.perf_counter()

    l = [int(s) for s in required_cuts.keys()]
    q = [int(s) for s in required_cuts.values()]

    reuse_pattern_counts = []
    all_combinations = []

    if reuse_rods:
        for rod_length, rod_count in reuse_rods.items():
            rod_combinations = generate_all_combinations([rod_length], required_cuts)
            pattern_count = len(rod_combinations)
            reuse_pattern_counts.append({
                'rod_length': rod_length,
                'count': rod_count,
                'pattern_count': pattern_count
            })
            all_combinations.extend(rod_combinations)

    normal_combinations = generate_all_combinations(available_rods, required_cuts)
    all_combinations.extend(normal_combinations)
    combinations_count = len(all_combinations)

    result_data = {
        "combinations_count": combinations_count,
        "available_rods": available_rods,
        "diameter": diameter,
        "uploaded_file": uploaded_file,
        "input_method": input_method,
        "reuse_rods": reuse_rods if reuse_rods else {},
    }

    if not all_combinations:
        result_data["error"] = "有効な組み合わせが見つかりませんでした"
        return result_data
    else:
        a = []
        c = []
        for combo in all_combinations:
            a.append([combo['cuts'].count(i) for i in l])
            c.append(combo['loss'])

        optimal_solution, optimal_value = optimal_cutting_plan(c, a, q, time_limit, reuse_pattern_counts)

        end_time = time.perf_counter()
        processing_time = end_time - start_time

        if optimal_solution is not None:
            total_rod_length = 0
            used_length = 0
            used_list = []
            cutting_patterns = []

            used_reuse_rods = {}
            if reuse_rods:
                for rod_length in reuse_rods.keys():
                    used_reuse_rods[rod_length] = 0

            for i in range(len(all_combinations)):
                if int(optimal_solution[i]) > 0:
                    pattern = {
                        'rod_length': all_combinations[i]['rod_length'],
                        'cuts': all_combinations[i]['cuts'],
                        'loss': all_combinations[i]['rod_length'] - sum(all_combinations[i]['cuts']),
                        'num': int(optimal_solution[i])
                    }
                    cutting_patterns.append(pattern)
                    total_rod_length += all_combinations[i]['rod_length'] * int(optimal_solution[i])
                    used_length += sum(all_combinations[i]['cuts']) * int(optimal_solution[i])
                    used_list.extend(all_combinations[i]['cuts'] * int(optimal_solution[i]))

                    if reuse_rods and all_combinations[i]['rod_length'] in reuse_rods:
                        used_reuse_rods[all_combinations[i]['rod_length']] += int(optimal_solution[i])

            remaining_reuse_rods = {}
            if reuse_rods:
                for rod_length, total_count in reuse_rods.items():
                    used_count = used_reuse_rods.get(rod_length, 0)
                    remaining_count = total_count - used_count
                    if remaining_count > 0:
                        remaining_reuse_rods[rod_length] = remaining_count

            used_count = [used_list.count(i) for i in l]

            if used_count == q:
                expanded_df, count_df, cutting_count_df, sheet_cutting_data_all, cutting_unique_values = create_result_sheets(
                    pd.DataFrame([
                        {
                            'id': i + 1,
                            'times': pattern['num'],
                            'loss': pattern['loss'],
                            'base': pattern['rod_length'],
                            'pattern': ','.join(map(str, pattern['cuts'])),
                        }
                        for i, pattern in enumerate(cutting_patterns)
                    ]), diameter, uploaded_file if input_method == "XLSXファイルアップロード" else None
                )

                # 切断パターンをシートに割り当て
                assigned_patterns, final_inventory = assign_patterns_to_sheets(
                    cutting_patterns,
                    sheet_cutting_data_all,
                    initial_inventory=reuse_rods if reuse_rods else None,
                )

                # 切断順序付き結果を作成
                df_results_list = []
                for i, pattern in enumerate(assigned_patterns):
                    operation = pattern.get('operation', '切断')

                    if operation == '切断':
                        df_results_list.append({
                            'id': i + 1,
                            '切断順序': pattern.get('cutting_order', i + 1),
                            'シート名': pattern.get('sheet_name', ''),
                            '操作': operation,
                            'base': pattern['rod_length'],
                            'pattern': ','.join(map(str, pattern['cuts'])),
                            'loss': pattern['loss'],
                        })
                    elif operation == '取出':
                        inventory_items = pattern.get('inventory_items', {})
                        items_str = ', '.join([f"{length}mm×{count}本" for length, count in sorted(inventory_items.items(), reverse=True)])
                        df_results_list.append({
                            'id': i + 1,
                            '切断順序': pattern.get('cutting_order', i + 1),
                            'シート名': pattern.get('sheet_name', ''),
                            '操作': operation,
                            'base': None,
                            'pattern': items_str,
                            'loss': 0,
                        })
                    elif operation == '保管':
                        inventory_items = pattern.get('inventory_items', {})
                        items_str = ', '.join([f"{length}mm×{count}本" for length, count in sorted(inventory_items.items(), reverse=True)])
                        df_results_list.append({
                            'id': i + 1,
                            '切断順序': pattern.get('cutting_order', i + 1),
                            'シート名': pattern.get('sheet_name', ''),
                            '操作': operation,
                            'base': None,
                            'pattern': items_str,
                            'loss': 0,
                        })

                df_results = pd.DataFrame(df_results_list)
                df_results['base'] = df_results['base'].astype('Int64')

                loss = total_rod_length - used_length
                yield_rate = used_length * 100 / total_rod_length

                result_data.update(
                    {
                        "success": True,
                        "processing_time": processing_time,
                        "total_rod_length": total_rod_length,
                        "used_length": used_length,
                        "cutting_patterns": cutting_patterns,
                        "df_results": df_results,
                        "expanded_df": expanded_df,
                        "count_df": count_df,
                        "cutting_count_df": cutting_count_df,
                        "sheet_cutting_data_all": sheet_cutting_data_all,
                        "cutting_unique_values": cutting_unique_values,
                        "remaining_reuse_rods": remaining_reuse_rods,
                        "reuse_rods": reuse_rods if reuse_rods else {},
                        "final_inventory": final_inventory,
                    }
                )
            else:
                result_data["success"] = False
                result_data["error"] = (
                    f"要求本数と切り出し個数が一致しません。要求: {q}, 切り出し: {used_count}"
                )
        else:
            result_data["success"] = False
            result_data["error"] = "最適解が見つかりませんでした"
    return result_data

def calculate_total_reuse_length(result, recalc_result):
    if not result.get("success"):
        return None

    remaining_reuse_rods = result.get("remaining_reuse_rods", {})
    scrap_above_threshold = recalc_result["scrap_above_threshold"]
    total_reuse_length = 0

    for item in scrap_above_threshold:
        total_reuse_length += item["length"] * item["count"]
    for length, count in remaining_reuse_rods.items():
        total_reuse_length += length * count

    return total_reuse_length

def display_optimization_results(result, scrap_threshold, tab_key=""):
    if not result.get("success"):
        st.error(result.get("error", "最適化に失敗しました"))
        return

    cutting_patterns = result["cutting_patterns"]
    processing_time = result["processing_time"]
    combinations_count = result["combinations_count"]
    df_results = result["df_results"]
    diameter = result["diameter"]
    expanded_df = result["expanded_df"]
    count_df = result["count_df"]
    cutting_count_df = result.get("cutting_count_df")
    cutting_unique_values = result.get("cutting_unique_values")

    recalc_result = recalculate_with_threshold(cutting_patterns, scrap_threshold)
    total_reuse_length = calculate_total_reuse_length(result, recalc_result)

    st.write(f"**切断パターン組み合わせ:** {combinations_count:,}")
    st.write("**処理時間:**", f"{processing_time:.2f} 秒")
    st.success("最適化が完了しました!")

    col_summary1, col_summary2, col_summary3 = st.columns([1, 1, 1])
    with col_summary1:
        st.metric("歩留り率(再利用なし)", f"{recalc_result['yield_rate']:.2f}%")
        st.metric("総材料長", f"{recalc_result['total_rod_length']:,} mm")
    with col_summary2:
        st.metric(
            "歩留り率(再利用あり)",
            f"{recalc_result['yield_rate_with_threshold']:.2f}%",
            delta=f"{recalc_result['yield_rate_with_threshold'] - recalc_result['yield_rate']:.2f}%",
        )
        st.metric("端材(再利用なし)", f"{recalc_result['loss']:,} mm")
    with col_summary3:
        st.metric("再利用端材の総長", f"{total_reuse_length:,} mm")
        st.metric("端材(再利用あり)", f"{recalc_result['scrap_below_threshold']:,} mm")

    st.write("切断パターン（シート割り当て順）")

    reuse_rods = result.get("reuse_rods", {})
    
    # シート名でグループ化して色分け表示
    def highlight_sheet_rows(row):
        if row['base'] in reuse_rods:
            return ['background-color: #ffff99'] * len(row)
        else:
            # シート名ごとに異なる背景色（薄い色）
            sheet_colors = {
                '未割り当て': 'background-color: #ffcccc',
            }
            sheet_name = row.get('シート名', '')
            if sheet_name and sheet_name != '未割り当て':
                # シート名のハッシュ値から色を生成
                color_idx = hash(sheet_name) % 5
                colors = ['#e6f2ff', '#ffe6f2', '#f2ffe6', '#fff2e6', '#f2e6ff']
                return [f'background-color: {colors[color_idx]}'] * len(row)
            elif sheet_name in sheet_colors:
                return [sheet_colors[sheet_name]] * len(row)
            else:
                return [''] * len(row)
    
    styled_df = df_results.style.apply(highlight_sheet_rows, axis=1)
    st.dataframe(styled_df, use_container_width=True)

    st.write("ダウンロード")
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        expanded_df.to_excel(writer, sheet_name="最適切断パターン", index=False)
        count_df.to_excel(writer, sheet_name="出力結果集計表", index=False)
        if cutting_count_df is not None:
            cutting_count_df.to_excel(writer, sheet_name="切断指示集計表", index=False)

    excel_buffer.seek(0)
    workbook = openpyxl.load_workbook(excel_buffer)

    if "出力結果集計表" in workbook.sheetnames:
        ws_output = workbook["出力結果集計表"]
        last_row = ws_output.max_row

        for col in range(2, ws_output.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_output[f"{col_letter}{last_row}"] = (
                f"=SUM({col_letter}2:{col_letter}{last_row-1})"
            )

    if cutting_count_df is not None and "切断指示集計表" in workbook.sheetnames:
        ws_cutting = workbook["切断指示集計表"]
        last_row_cutting = ws_cutting.max_row

        for col in range(2, ws_cutting.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_cutting[f"{col_letter}{last_row_cutting}"] = (
                f"=SUM({col_letter}2:{col_letter}{last_row_cutting-1})"
            )

        ws_summary = workbook.create_sheet("サマリー")
        ws_summary["A1"] = "径"
        ws_summary["B1"] = diameter

        ws_summary["A2"] = "歩留り率(再利用なし)"
        ws_summary["B2"] = recalc_result["yield_rate"]
        ws_summary["B2"].number_format = "0.00%"
        ws_summary["B2"].value = recalc_result["yield_rate"] / 100

        ws_summary["A3"] = "端材(再利用なし)[mm]"
        ws_summary["B3"] = recalc_result["loss"]
        ws_summary["B3"].number_format = "#,##0"

        ws_summary["A4"] = "端材閾値"
        ws_summary["B4"] = scrap_threshold

        ws_summary["A5"] = "歩留り率(再利用あり)"
        ws_summary["B5"] = recalc_result["yield_rate_with_threshold"]
        ws_summary["B5"].number_format = "0.00%"
        ws_summary["B5"].value = recalc_result["yield_rate_with_threshold"] / 100

        ws_summary["A6"] = "端材(再利用あり)[mm]"
        ws_summary["B6"] = recalc_result["scrap_below_threshold"]
        ws_summary["B6"].number_format = "#,##0"

        ws_summary["A7"] = "総材料長"
        ws_summary["B7"] = recalc_result["total_rod_length"]
        ws_summary["B7"].number_format = "#,##0"

        ws_summary["A8"] = "処理時間(s)"
        ws_summary["B8"] = processing_time
        ws_summary["B8"].number_format = "#,##0.00"

        if cutting_unique_values:
            ws_summary["A10"] = "差分(出力結果 - 切断指示)"

            ws_summary["A11"] = "長さ(mm)"
            ws_summary["B11"] = "出力結果"
            ws_summary["C11"] = "切断指示"
            ws_summary["D11"] = "差分"

            row_num = 12
            for i, value in enumerate(cutting_unique_values):
                col_letter_cutting = openpyxl.utils.get_column_letter(i + 2)
                col_letter_output = openpyxl.utils.get_column_letter(i + 2)

                ws_summary[f"A{row_num}"] = str(value)
                ws_summary[f"B{row_num}"] = (
                    f"=出力結果集計表!{col_letter_output}{last_row}"
                )
                ws_summary[f"C{row_num}"] = (
                    f"=切断指示集計表!{col_letter_cutting}{last_row_cutting}"
                )
                ws_summary[f"D{row_num}"] = f"=B{row_num}-C{row_num}"
                row_num += 1

        new_buffer = io.BytesIO()
        workbook.save(new_buffer)
        workbook.close()
        new_buffer.seek(0)

        st.download_button(
            label="最適化結果シート",
            data=new_buffer.getvalue(),
            file_name=f"result_sheet_{diameter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"excel_{tab_key}_{scrap_threshold}",
        )

def display_consolidated_reusable_scrap(optimization_results, scrap_threshold):
    """
    全径統合の再利用端材リストを表示・ダウンロード
    
    Parameters:
    - optimization_results: 径ごとの最適化結果の辞書
    - scrap_threshold: 端材閾値
    """
    st.write("---")
    st.write("### 📥 全径統合 再利用端材リスト")
    
    # 全径の最終在庫と端材を統合
    all_reusable_scrap = []
    
    for diameter, result in optimization_results.items():
        if not result.get("success"):
            continue
        
        # 最終在庫
        final_inventory = result.get("final_inventory", {})
        for length, count in final_inventory.items():
            all_reusable_scrap.append({
                '鉄筋径': diameter,
                '長さ (mm)': length,
                '本数': count
            })
        
        # 端材（閾値以上）
        recalc_result = recalculate_with_threshold(result["cutting_patterns"], scrap_threshold)
        for item in recalc_result["scrap_above_threshold"]:
            all_reusable_scrap.append({
                '鉄筋径': diameter,
                '長さ (mm)': item['length'],
                '本数': item['count']
            })
    
    if all_reusable_scrap:
        df_all_reusable = pd.DataFrame(all_reusable_scrap)
        
        # 同じ径・長さをグループ化
        df_all_reusable = df_all_reusable.groupby(['鉄筋径', '長さ (mm)'], as_index=False)['本数'].sum()
        df_all_reusable = df_all_reusable.sort_values(['鉄筋径', '長さ (mm)'], ascending=[True, False])
        
        st.dataframe(df_all_reusable, use_container_width=True)
        
        csv_all_reusable = df_all_reusable.to_csv(index=False, encoding="utf-8-sig")
        
        st.download_button(
            label=f"全径統合 再利用端材リスト (≥{scrap_threshold}mm)",
            data=csv_all_reusable,
            file_name=f"reusable_scrap_all_diameters_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key=f"reuse_all_diameters_{scrap_threshold}",
        )
    else:
        st.info(f"再利用可能な端材(≥{scrap_threshold}mm)はありません")

def main():
    st.set_page_config(page_title="鉄筋切断最適化アプリ", layout="wide")

    # カスタムCSSで行間を詰める
    st.markdown("""
        <style>
        .stCheckbox {
            margin-bottom: -20px;
        }
        div[data-testid="column"] {
            padding-top: 0px;
            padding-bottom: 0px;
        }
        </style>
        """, unsafe_allow_html=True)

    st.title("🔧 鉄筋切断最適化アプリ")
    st.write("鉄筋の切断パターンを最適化し、廃材を最小化します。")

    # サイドバーに使用方法を表示
    with st.sidebar:
        st.header("📖 使用方法")
        st.markdown("""
        ### 使用手順
        1. **材料長さ選択**: 使用する鉄筋径と材料長さをチェックボックスで選択
        2. **時間設定**: 最適化計算の制限時間を設定（複雑な問題ほど長めに設定）
        3. **切断指示**: XLSXファイルをアップロード（全鉄筋径のデータを自動読み込み）
        4. **再利用端材（オプション）**: 既存の端材がある場合はCSVファイルをアップロード
        5. **端材閾値**: 再利用可能な端材の最小長さを設定
        6. **最適化実行**: ボタンをクリックして最適化を実行
        7. **結果確認**: 径ごとのタブで歩留り率や切断パターンを確認
        8. **ダウンロード**: 再利用端材リストと最適化結果シートをダウンロード

        ### ファイル形式
        - **XLSXファイル**: 切断集計表（「鉄筋径」列を含む形式）
        - **CSVファイル**: 鉄筋径, 端材の長さ (mm), 本数 の3列形式
        """)

    col1, col2 = st.columns([1, 1])

    with col1:
        st.header("⚙️ 設定")

        # 変数の初期化
        uploaded_file = None
        reuse_rods = {}
        st.write("使用する材料長さをチェックしてください：")

        # session_stateの初期化
        if 'selected_materials' not in st.session_state:
            st.session_state.selected_materials = {}

        # 各長さごとにチェックボックスを横に並べて表示
        for idx, length in enumerate(ALL_AVAILABLE_LENGTHS):
            is_last = (idx == len(ALL_AVAILABLE_LENGTHS) - 1)            
            cols = st.columns(len(DIAMETERS))

            for i, diameter in enumerate(DIAMETERS):
                with cols[i]:
                    is_checked = st.checkbox(
                        f"{diameter}-{length}",
                        value=False,
                        key=f"material_{diameter}_{length}"
                    )

                    # 選択状態を保存
                    if diameter not in st.session_state.selected_materials:
                        st.session_state.selected_materials[diameter] = []

                    if is_checked and length not in st.session_state.selected_materials[diameter]:
                        st.session_state.selected_materials[diameter].append(length)
                    elif not is_checked and length in st.session_state.selected_materials[diameter]:
                        st.session_state.selected_materials[diameter].remove(length)

            # 最後の行の場合は通常の余白を追加
            if is_last:
                st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)

        selected_diameters = []
        for diameter in DIAMETERS:
            lengths = sorted(st.session_state.selected_materials.get(diameter, []))
            if lengths:
                selected_diameters.append(diameter)

        if not selected_diameters:
            st.warning("材料長さを選択してください")

        st.write("---")

        time_limit = st.number_input(
            "最適化の制限時間 (10~3600 秒)",
            min_value=10,
            max_value=3600,
            value=120,
            step=10,
            help="最適化計算の制限時間を設定します。大きな問題では時間を長く設定することを推奨します。"
        )
        st.write(f"現在の設定: {time_limit}秒")

        st.subheader("XLSXファイルアップロード")

        uploaded_file = st.file_uploader(
            "XLSXファイルを選択してください",
            type=['xlsx'],
            help="切断集計表のXLSXファイルをアップロードしてください"
        )

        if uploaded_file is not None:
            with st.spinner("ファイルを解析中..."):
                # 全ての鉄筋径についてデータを読み込む
                all_cutting_data = {}
                for diameter in DIAMETERS:
                    cutting_data = read_cutting_data_from_xlsx(uploaded_file, diameter)
                    if cutting_data:
                        all_cutting_data[diameter] = cutting_data

            if all_cutting_data:
                st.success(f"XLSXファイルを正常に読み込みました。{len(all_cutting_data)}種類の径のデータが見つかりました。")

                # 読み込まれたデータを径ごとに表示
                for diameter in DIAMETERS:
                    if diameter in all_cutting_data:
                        with st.expander(f"**{diameter}の切断指示データ**"):
                            df_preview = pd.DataFrame([
                                {'長さ (mm)': length, '本数': count}
                                for length, count in sorted(all_cutting_data[diameter].items(), reverse=True)
                            ])
                            st.dataframe(df_preview, use_container_width=True)

                            total_pieces = sum(all_cutting_data[diameter].values())
                            cutting_types = len(all_cutting_data[diameter])
                            st.write(f"切断種類数: {cutting_types}種類, 総切断本数: {total_pieces:,}本")

                # session_stateに保存
                st.session_state.xlsx_cutting_data = all_cutting_data
            else:
                st.error("XLSXファイルから有効なデータが見つかりませんでした。")
                if 'xlsx_cutting_data' in st.session_state:
                    del st.session_state.xlsx_cutting_data
        else:
            # ファイルがアップロードされていない場合はクリア
            if 'xlsx_cutting_data' in st.session_state:
                del st.session_state.xlsx_cutting_data

        st.subheader("再利用端材 (オプション)")
        scrap_csv_file = st.file_uploader(
            "再利用端材のCSVファイル (オプション)",
            type=["csv"],
            help="鉄筋径,長さ (mm),本数 の形式のCSVファイルをアップロードしてください",
            key="scrap_csv_uploader",
        )

        if scrap_csv_file is not None:
            with st.spinner("端材CSVを解析中..."):
                scrap_data_by_diameter = read_scrap_data_from_csv(scrap_csv_file)

            if scrap_data_by_diameter:
                st.success(f"再利用端材のCSVファイルを正常に読み込みました。")

                st.write("**読み込まれた再利用端材:**")

                # 径ごとに表示
                for diameter in DIAMETERS:
                    if diameter in scrap_data_by_diameter:
                        with st.expander(f"**{diameter}の再利用端材**"):
                            df_scrap_preview = pd.DataFrame(
                                [
                                    {"長さ (mm)": length, "本数": count}
                                    for length, count in sorted(
                                        scrap_data_by_diameter[diameter].items(), reverse=True
                                    )
                                ]
                            )
                            st.dataframe(df_scrap_preview, use_container_width=True)

                            total_scrap_pieces = sum(scrap_data_by_diameter[diameter].values())
                            scrap_types = len(scrap_data_by_diameter[diameter])
                            st.write(
                                f"端材種類数: {scrap_types}種類, 総端材本数: {total_scrap_pieces:,}本"
                            )

                # session_stateに保存
                st.session_state.scrap_data_by_diameter = scrap_data_by_diameter
            else:
                st.error("再利用端材のCSVファイルの読み込みに失敗しました。")
                if "scrap_data_by_diameter" in st.session_state:
                    del st.session_state.scrap_data_by_diameter
        else:
            if "scrap_data_by_diameter" in st.session_state:
                del st.session_state.scrap_data_by_diameter

    with col2:
        st.header("🎯 最適化結果")

        scrap_threshold = st.number_input(
            "端材閾値 (mm)",
            min_value=0,
            max_value=2000,
            value=400,
            step=50,
            help="この閾値未満の端材のみを廃材として歩留り率を計算します。閾値以上の端材は再利用可能として扱われます。"
        )
        st.write(f"現在の設定: {scrap_threshold}mm未満を廃材として扱う")

        if 'optimization_results' not in st.session_state:
            st.session_state.optimization_results = {}

        # 最適化実行の準備
        can_execute = False
        execution_data = {}

        if 'xlsx_cutting_data' in st.session_state and st.session_state.xlsx_cutting_data:
            can_execute = True
            execution_data = st.session_state.xlsx_cutting_data

        if st.button("最適化を実行", type="primary") and can_execute:
            st.session_state.optimization_results = {}
            st.session_state.optimization_attempted = True  # 実行試行フラグ

            # チェックボックスで選択されている径を取得
            selected_diameters = [d for d in DIAMETERS if st.session_state.selected_materials.get(d, [])]

            if not selected_diameters:
                st.error("材料長さを選択してください")
            else:
                # バリデーションを実行
                is_valid, error_messages = validate_material_selection(
                    execution_data, 
                    st.session_state.selected_materials
                )

                if not is_valid:
                    st.error("**選択された材料長に問題があります：**")
                    for msg in error_messages:
                        st.warning(msg)
                else:
                    # 各径で最適化を実行
                    scrap_data_by_diameter = st.session_state.get('scrap_data_by_diameter', {})

                    for diameter in selected_diameters:
                        available_rods = sorted(st.session_state.selected_materials.get(diameter, []))

                        if not available_rods:
                            st.warning(f"{diameter}の材料長さが選択されていません")
                            continue

                        # 切断指示データを取得
                        if diameter not in execution_data:
                            st.warning(f"{diameter}の切断指示データがありません")
                            continue

                        diameter_required_cuts = execution_data[diameter]

                        # この径の再利用端材を取得
                        diameter_reuse_rods = scrap_data_by_diameter.get(diameter, {})

                        with st.spinner(f"{diameter} の最適化を実行中..."):
                            result = execute_optimizer(
                                available_rods,
                                diameter_required_cuts,
                                diameter,
                                time_limit,
                                uploaded_file,
                                "XLSXファイルアップロード",
                                scrap_threshold,
                                diameter_reuse_rods if diameter_reuse_rods else None,
                            )
                            st.session_state.optimization_results[diameter] = result

        if st.session_state.optimization_results:
            # 結果がある径のみを取得
            result_diameters = list(st.session_state.optimization_results.keys())

            if len(result_diameters) == 0:
                st.warning("最適化結果がありません")
            else:
                # シート単位での統合結果を表示
                st.write("### 📋 シート別切断指示（全径統合）")

                consolidated_df, sheet_order, sheet_color_map = consolidate_results_by_sheet(st.session_state.optimization_results)

                if not consolidated_df.empty:
                    # シート名で色分け表示（操作によらず同じシートは同じ色）
                    def highlight_sheet_rows_consolidated(row):
                        sheet_name = row.get('シート名', '')
                        color = sheet_color_map.get(sheet_name, '#ffffff')
                        return [f'background-color: {color}'] * len(row)

                    # 表示用のDataFrameを作成
                    display_df = consolidated_df.copy()
                    display_df['id'] = display_df['統合切断順序']

                    # 列の順序を整理
                    column_order = ['id', 'シート名', '径', '操作', '本数', 'base', 'pattern', 'loss']
                    display_df = display_df[[col for col in column_order if col in display_df.columns]]

                    styled_df = display_df.style.apply(highlight_sheet_rows_consolidated, axis=1)
                    st.dataframe(styled_df, use_container_width=True, hide_index=True)

                    # シート別切断順序のダウンロード
                    st.write("---")
                    st.write("**統合結果ダウンロード:**")

                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                        # シート別切断順序
                        display_df.to_excel(writer, sheet_name="シート別切断順序", index=False)

                    excel_buffer.seek(0)

                    # openpyxlで色分けを追加
                    workbook = openpyxl.load_workbook(excel_buffer)
                    ws = workbook["シート別切断順序"]

                    # ヘッダー行を特定（1行目）
                    header_row = 1

                    # シート名列のインデックスを取得
                    sheet_name_col_idx = None
                    for col_idx, cell in enumerate(ws[header_row], start=1):
                        if cell.value == 'シート名':
                            sheet_name_col_idx = col_idx
                            break

                    if sheet_name_col_idx:
                        # データ行に色を適用（2行目以降）
                        for row_idx in range(header_row + 1, ws.max_row + 1):
                            sheet_name_cell = ws.cell(row=row_idx, column=sheet_name_col_idx)
                            sheet_name = sheet_name_cell.value

                            if sheet_name and sheet_name in sheet_color_map:
                                # 16進数カラーコードをRGB形式に変換（#を除く）
                                color_hex = sheet_color_map[sheet_name].replace('#', '')
                                fill = openpyxl.styles.PatternFill(
                                    start_color=color_hex,
                                    end_color=color_hex,
                                    fill_type='solid'
                                )

                                # その行の全セルに背景色を適用
                                for col_idx in range(1, ws.max_column + 1):
                                    ws.cell(row=row_idx, column=col_idx).fill = fill

                    # 修正したworkbookを新しいバッファに保存
                    new_buffer = io.BytesIO()
                    workbook.save(new_buffer)
                    workbook.close()
                    new_buffer.seek(0)

                    st.download_button(
                        label="シート別切断順序をダウンロード",
                        data=new_buffer.getvalue(),
                        file_name=f"cutting_order_by_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="consolidated_download"
                    )

                st.write("---")
                st.write("### 📊 径別詳細結果")

                # 径別の詳細結果はタブで表示
                if len(result_diameters) == 1:
                    # 1つの径のみの場合はタブを作らず直接表示
                    diameter = result_diameters[0]
                    result = st.session_state.optimization_results[diameter]

                    st.write(f"**{diameter} の最適化結果**")
                    display_optimization_results(
                        result,
                        scrap_threshold,
                        diameter
                    )
                else:
                    # 複数の径がある場合はタブで表示
                    diameter_tabs = st.tabs(result_diameters)

                    for tab_idx, diameter in enumerate(result_diameters):
                        with diameter_tabs[tab_idx]:
                            result = st.session_state.optimization_results[diameter]
                            display_optimization_results(
                                result,
                                scrap_threshold,
                                diameter
                            )
                # 全径統合の再利用端材リストを表示
                display_consolidated_reusable_scrap(st.session_state.optimization_results, scrap_threshold)
        elif not can_execute:
            st.info("XLSXファイルをアップロードしてください。")
        elif not st.session_state.get('optimization_attempted', False):
            # 最適化が一度も試行されていない場合のみメッセージ表示
            st.info("最適化を実行してください。")

if __name__ == "__main__":
    main()
