import openpyxl
from collections import defaultdict

def find_cell_position(worksheet, search_text):
    """
    ワークシート内で指定されたテキストを含むセルの位置を検索
    
    Args:
        worksheet: openpyxlのワークシートオブジェクト
        search_text: 検索するテキスト
    
    Returns:
        tuple: (行番号, 列番号) または None（見つからない場合）
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value and search_text in str(cell.value):
                return cell.row, cell.column
    return None

def get_diameter_column_index(worksheet, base_row, base_col, target_diameter):
    """
    指定された径の列インデックスを取得
    
    Args:
        worksheet: openpyxlのワークシートオブジェクト
        base_row: 基準行（鉄筋径が記載されている行）
        base_col: 基準列
        target_diameter: 対象の径（例: "D10", "D13"）
    
    Returns:
        int: 径の列インデックス（1ベース）または None
    """
    # 基準行を右に検索して径を探す
    col = base_col + 1
    max_col = worksheet.max_column
    
    while col <= max_col:
        cell_value = worksheet.cell(row=base_row, column=col).value
        if cell_value and str(cell_value).strip() == target_diameter:
            return col
        col += 1
    
    return None

def extract_cutting_data_from_sheet(worksheet, target_diameter):
    """
    単一シートから指定された径の切断データを抽出
    
    Args:
        worksheet: openpyxlのワークシートオブジェクト
        target_diameter: 対象の径（例: "D10", "D13"）
    
    Returns:
        dict: {長さ: 本数} の辞書
    """
    # "鉄筋径"を検索
    base_position = find_cell_position(worksheet, "鉄筋径")
    if not base_position:
        return {}
    
    base_row, base_col = base_position
    
    # 指定された径の列インデックスを取得
    diameter_col = get_diameter_column_index(worksheet, base_row, base_col, target_diameter)
    if not diameter_col:
        return {}
    
    # 長さと本数のデータを抽出
    cutting_data = {}
    row = base_row + 1  # 鉄筋径の次の行から開始
    max_row = worksheet.max_row
    
    while row <= max_row:
        # 長さを取得（基準列の下）
        length_cell = worksheet.cell(row=row, column=base_col+1)
        # 本数を取得（径の列）
        count_cell = worksheet.cell(row=row, column=diameter_col)
        
        # 両方の値が存在し、有効な数値の場合のみ追加
        if (length_cell.value is not None and count_cell.value is not None and
            isinstance(length_cell.value, (int, float)) and 
            isinstance(count_cell.value, (int, float)) and
            count_cell.value > 0):
            
            length = int(length_cell.value)
            count = int(count_cell.value)
            cutting_data[length] = count
        
        # 長さが記載されていない行が続いたら終了
        if length_cell.value is None or length_cell.value == '':
            # 連続する空セルが3行続いたら終了
            empty_count = 0
            temp_row = row
            while temp_row <= min(row + 2, max_row):
                if worksheet.cell(row=temp_row, column=base_col).value is None:
                    empty_count += 1
                temp_row += 1
            if empty_count >= 3:
                break
        
        row += 1
    
    return cutting_data

def read_cutting_data(file_path, target_diameter):
    """
    エクセルファイルから指定された径の切断データを全シートから読み込み、合算
    
    Args:
        file_path: エクセルファイルのパス
        target_diameter: 対象の径（例: "D10", "D13"）
    
    Returns:
        dict: 全シートの合算された {長さ: 本数} の辞書
    """
    try:
        # エクセルファイルを開く
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 合算用の辞書
        total_cutting_data = defaultdict(int)
        
        # 各シートを処理
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # シートから切断データを抽出
            sheet_cutting_data = extract_cutting_data_from_sheet(worksheet, target_diameter)
            
            # デバッグ用出力
            # if sheet_cutting_data:
            #     print(f"'{sheet_name}' - {target_diameter}: {sheet_cutting_data}")
            
            # 合算
            for length, count in sheet_cutting_data.items():
                total_cutting_data[length] += count
        
        # defaultdictを通常の辞書に変換
        result = dict(total_cutting_data)
        
        workbook.close()
        return result
        
    except FileNotFoundError:
        print(f"ファイルが見つかりません: {file_path}")
        return {}
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        return {}

def get_patterns_from_xlsx(file_path, diameter):
    """
    エクセルファイルから指定された径のパターンを取得
    cutting_optimizer_ver1.pyのgetPatterns関数と同様のインターフェース
    
    Args:
        file_path: エクセルファイルのパス
        diameter: 径（例: "D10", "D13"）
    
    Returns:
        dict: {'base_patterns': [利用可能な棒の長さのリスト], 'tasks': {長さ: 本数}}
    """
    # 各径の原材料の長さ（cutting_optimizer_ver2.pyのBASE_PATTERNSと同じ）
    BASE_PATTERNS = {
        'D10': [4000, 4500, 5500, 6000],
        'D13': [3000, 4000, 4500, 5500, 6000, 7500],
        'D16': [4000, 4500, 5500, 6000, 7000],
        'D19': [3500, 4000, 4500, 5500, 6000],
        'D22': [4000, 4500, 5500, 6000]
    }
    
    # 切断データを読み込み
    cutting_data = read_cutting_data(file_path, diameter)
    
    # 結果を返す
    result = {
        'base_patterns': BASE_PATTERNS[diameter],
        'tasks': cutting_data
    }
    
    return result

def main(diameter):
    """
    メイン処理（テスト用）
    """
    file_path = '切断集計表.xlsx'
    
    # テスト用：鉄筋径リスト
    # diameters = ['D10', 'D13', 'D16', 'D19', 'D22']
    
    # for diameter in diameters:
    print(f"\n=== {diameter} ===")
    cutting_data = read_cutting_data(file_path, diameter)
    
    if cutting_data:
        print(f"合算結果: {cutting_data}")
        
        # 統計情報
        total_pieces = sum(cutting_data.values())
        print(f"総本数: {total_pieces}本")
        print(f"長さの種類: {len(cutting_data)}種類")
    else:
        print("データが見つかりませんでした")

if __name__ == "__main__":
    diameter = 'D10'
    main(diameter)